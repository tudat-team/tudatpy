#!/usr/bin/env python3
"""
Serialization Compliance Tracker

Scans the Tudat codebase and produces an Excel report showing:
  1. What "Settings" classes exist
  2. What classes implement load/save for serialization
  3. What classes implement operator==
  4. What classes have C++ roundtrip testing
  5. What classes have Python exposure for: Equality, Pickle, save_to_*
  6. What classes have Python roundtrip testing

Usage:
    python track_serialization.py [--csv] [--output PATH]

Output: track_serialization.xlsx (default) or CSV if --csv is given.
"""

import re
import argparse
from pathlib import Path
from collections import defaultdict
from dataclasses import dataclass, field

# ── workspace root ──────────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parent
INCLUDE_DIR = ROOT / "include" / "tudat"
SRC_TUDATPY_DIR = ROOT / "src" / "tudatpy"
TESTS_CPP_DIR = ROOT / "tests" / "test_tudat" / "src" / "io"
TESTS_PY_DIR = ROOT / "tests" / "test_tudatpy"

DEFAULT_EXCEL = ROOT / "track_serialization.xlsx"


# ── data structures ─────────────────────────────────────────────────────────


@dataclass
class ClassInfo:
    name: str
    is_settings: bool = False
    header_file: str = ""
    has_save_load: bool = False
    has_operator_eq: bool = False
    has_equals_method: bool = False
    has_file_io_binary: bool = False  # declares/defines saveToBinary + loadFromBinary
    has_file_io_json: bool = False  # declares/defines saveToJson + loadFromJson
    file_io_polymorphic: bool = False  # load returns shared_ptr<Base> (polymorphic)
    base_classes: list = field(default_factory=list)  # direct base class names
    cpp_test_files: list = field(default_factory=list)
    py_expose_file: str = ""
    py_has_equals: bool = False
    py_has_pickle: bool = False
    py_has_save_to: bool = False
    py_factory_names: list = field(default_factory=list)  # Python factory function names
    py_test_files: list = field(default_factory=list)


# ── helpers ──────────────────────────────────────────────────────────────────


def strip_template(name: str) -> str:
    """Strip template params and namespace: 'Foo<T>' or 'ns::Foo' → 'Foo'"""
    base = name.split("<")[0].strip()
    if "::" in base:
        base = base.split("::")[-1]
    return base


def is_settings_class(name: str) -> bool:
    return name.lower().endswith("settings") and name != "Settings"


# ── C++ class name → Python factory function name mapping ─────────────────
#
# Many C++ serializable classes are instantiated in Python tests only through
# factory functions whose names don't match the C++ class name.  This lookup
# bridges the gap so the tracker can detect coverage of those classes.
#
# Key:
#   C++ class name (exact) → list of Python factory / submodule names that
#   create instances of that class and may appear in Python test code.

CPP_TO_PYTHON_FACTORY: dict[str, list[str]] = {
    # -- Acceleration settings (dynamics.propagation_setup.acceleration) --
    "DirectTidalDissipationAccelerationSettings": ["direct_tidal_dissipation_acceleration"],
    "EmpiricalAccelerationSettings": ["empirical"],
    "MomentumWheelDesaturationAccelerationSettings": ["momentum_wheel_desaturation_acceleration"],
    "MutualSphericalHarmonicAccelerationSettings": ["mutual_spherical_harmonic_gravity"],
    "RadiationPressureAccelerationSettings": ["radiation_pressure"],
    "RelativisticAccelerationCorrectionSettings": ["relativistic_correction"],
    "RTGAccelerationSettings": ["rtg"],
    "SphericalHarmonicAccelerationSettings": ["spherical_harmonic_gravity"],
    "ThrustAccelerationSettings": [
        "thrust_from_engines",
        "thrust_from_engine",
        "thrust_from_all_engines",
    ],
    "YarkovskyAccelerationSettings": ["yarkovsky"],
    # -- Torque settings (dynamics.propagation_setup.torque) --
    "SphericalHarmonicTorqueSettings": ["spherical_harmonic_gravitational"],
    # -- Propagation termination (dynamics.propagation_setup.propagator) --
    "PropagationTimeTerminationSettings": [
        "time_termination",
        "propagation_time_termination_settings",
    ],
    "PropagationCPUTimeTerminationSettings": [
        "cpu_time_termination",
        "propagation_cpu_time_termination_settings",
    ],
    "PropagationDependentVariableTerminationSettings": [
        "dependent_variable_termination",
        "propagation_dependent_variable_termination_settings",
    ],
    "PropagationHybridTerminationSettings": ["hybrid_termination"],
    "NonSequentialPropagationTerminationSettings": [
        "non_sequential_termination",
        "non_sequential_propagation_termination_settings",
    ],
    # -- Dependent variable settings (dynamics.propagation_setup.dependent_variable) --
    "SingleAccelerationDependentVariableSaveSettings": ["single_acceleration"],
    "SingleTorqueDependentVariableSaveSettings": ["single_torque"],
    "SphericalHarmonicAccelerationTermsDependentVariableSaveSettings": [
        "spherical_harmonic_terms_acceleration"
    ],
    "BodyAerodynamicAngleVariableSaveSettings": [
        "angle_of_attack",
        "sideslip_angle",
        "bank_angle",
        "heading_angle",
        "flight_path_angle",
    ],
    "ControlSurfaceCoefficientDependentVariableSettings": ["control_surface_deflection"],
    "IntermediateAerodynamicRotationVariableSaveSettings": [
        "intermediate_aerodynamic_rotation_matrix_variable"
    ],
    "LocalWindVelocityDependentVariableSaveSettings": ["local_wind_velocity"],
    "CrossSectionDependentVariableSaveSettings": ["actual_cross_section"],
    "IlluminatedPanelFractionDependentVariableSaveSettings": ["illuminated_panel_fraction"],
    "MinimumConstellationDistanceDependentVariableSaveSettings": ["minimum_body_distance"],
    "MinimumConstellationStationDistanceDependentVariableSaveSettings": [
        "minimum_visible_station_body_distances"
    ],
    "AccelerationPartialWrtStateSaveSettings": [
        "acceleration_partial_wrt_body_translational_state"
    ],
    "TotalAccelerationPartialWrtStateSaveSettings": [
        "total_acceleration_partial_wrt_body_translational_state"
    ],
    "GravityFieldVariationSettings": ["single_gravity_field_variation_acceleration"],
    "SingleVariationSphericalHarmonicAccelerationSaveSettings": [
        "single_gravity_field_variation_acceleration"
    ],
    "SingleVariationSingleTermSphericalHarmonicAccelerationSaveSettings": [
        "single_per_term_gravity_field_variation_acceleration"
    ],
    "TotalGravityFieldVariationSettings": ["total_gravity_field_variation_acceleration"],
    # -- Gravity field variation (dynamics.environment_setup.gravity_field_variation) --
    "BasicSolidBodyGravityFieldVariationSettings": [
        "solid_body_tide",
        "solid_body_tide_degree_variable_k",
        "solid_body_tide_degree_order_variable_k",
        "solid_body_tide_degree_variable_complex_k",
        "solid_multi_body_tide_degree_order_variable_k",
    ],
    "ModeCoupledSolidBodyGravityFieldVariationSettings": ["mode_coupled_solid_body_tide"],
    "PeriodicGravityFieldVariationsSettings": ["single_period_periodic", "periodic"],
    "PolynomialGravityFieldVariationsSettings": ["single_power_polynomial", "polynomial"],
    "TabulatedGravityFieldVariationSettings": ["tabulated"],
    # -- Observation dependent variables (estimation.observations_setup.observations_dependent_variables) --
    "StationAngleObservationDependentVariableSettings": [
        "elevation_angle_dependent_variable",
        "azimuth_angle_dependent_variable",
    ],
    "InterlinkObservationDependentVariableSettings": [
        "integration_time_dependent_variable",
        "link_end_epochs_dependent_variable",
    ],
    "AncillaryObservationDependentVariableSettings": ["retransmission_delays_dependent_variable"],
    "LightTimeCorrectionComponentsDependentVariableSettings": [
        "light_time_correction_components_dependent_variable"
    ],
    # -- Root finders (math.root_finders) --
    "RootFinderSettings": ["bisection", "newton_raphson", "halley", "secant"],
    # -- Ancillary simulation settings (estimation.observations_setup.ancillary_settings) --
    "ObservationAncillarySimulationSettings": ["doppler_ancillary_settings"],
}


# ── Phase 1: Scan headers for class inventory ───────────────────────────────


def scan_headers() -> dict[str, ClassInfo]:
    """Build inventory of all interesting classes from headers.

    For each header file, find all class declarations, then check
    per-class whether the class body contains save/load pairs and
    operator==, and which classes are settings classes.
    """
    results: dict[str, ClassInfo] = {}

    # Patterns for class/struct bodies: capture class name and position of '{'
    class_start_re = re.compile(
        r"(?:^|\n)\s*(?:class|struct)\s+(\w+)\s+[^{;]{0,2000}(\{)", re.MULTILINE
    )
    save_re = re.compile(r"void\s+save\s*\(\s*[^)]*Archive[^)]*\)")
    load_re = re.compile(r"void\s+load\s*\(\s*[^)]*Archive[^)]*\)")
    operator_eq_re = re.compile(r"(?:friend\s+)?(?:bool|auto)\s+operator\s*==\s*\([^)]*\)")
    equals_method_re = re.compile(r"bool\s+equals\s*\([^)]*\)\s*const(?:\s+override)?")
    # File IO patterns — match both the explicit declaration and the macro invocation.
    # The macros use __VA_ARGS__ so they can appear as:
    #   TUDAT_DEFINE_FILE_IO( ClassName )
    #   TUDAT_DEFINE_BINARY_IO( ClassName )
    #   TUDAT_DEFINE_JSON_IO( ClassName )
    #   TUDAT_DEFINE_FILE_IO_POLYMORPHIC( BaseName )
    #   TUDAT_DEFINE_JSON_IO_POLYMORPHIC( BaseName )
    #   TUDAT_DEFINE_BINARY_IO_POLYMORPHIC( BaseName )
    save_to_binary_re = re.compile(
        r"(?:"
        r"saveToBinary\s*\(\s*(?:const\s+)?std::string\s*&?\s*(?:const\s+)?(?:&\s*)?path\s*\)"
        r"|"
        r"TUDAT_DEFINE_(?:FILE_IO(?:_POLYMORPHIC)?|BINARY_IO(?:_POLYMORPHIC)?)\s*\("
        r")"
    )
    load_from_binary_re = re.compile(
        r"(?:"
        r"loadFromBinary\s*\(\s*(?:const\s+)?std::string\s*&?\s*(?:const\s+)?(?:&\s*)?path\s*\)"
        r"|"
        r"TUDAT_DEFINE_(?:FILE_IO(?:_POLYMORPHIC)?|BINARY_IO(?:_POLYMORPHIC)?)\s*\("
        r")"
    )
    save_to_json_re = re.compile(
        r"(?:"
        r"saveToJson\s*\(\s*(?:const\s+)?std::string\s*&?\s*(?:const\s+)?(?:&\s*)?path\s*\)"
        r"|"
        r"TUDAT_DEFINE_(?:FILE_IO(?:_POLYMORPHIC)?|JSON_IO(?:_POLYMORPHIC)?)\s*\("
        r")"
    )
    load_from_json_re = re.compile(
        r"(?:"
        r"loadFromJson\s*\(\s*(?:const\s+)?std::string\s*&?\s*(?:const\s+)?(?:&\s*)?path\s*\)"
        r"|"
        r"TUDAT_DEFINE_(?:FILE_IO(?:_POLYMORPHIC)?|JSON_IO(?:_POLYMORPHIC)?)\s*\("
        r")"
    )
    # Polymorphic load detection: returns shared_ptr<...> or uses _POLYMORPHIC macro variant
    poly_load_re = re.compile(
        r"static\s+std::shared_ptr\s*<"
        r"|"
        r"TUDAT_DEFINE_(?:FILE_IO_POLYMORPHIC|JSON_IO_POLYMORPHIC|BINARY_IO_POLYMORPHIC)\s*\("
    )

    for hdr in sorted(INCLUDE_DIR.rglob("*.h")):
        rel = str(hdr.relative_to(ROOT))
        try:
            content = hdr.read_text(errors="ignore")
        except Exception:
            continue

        # Find each class/struct with its opening brace position
        # Build list of (cls_name, body_start, body_end) where body
        # is the text between the opening '{' and matching '}'.
        classes = []
        for m in class_start_re.finditer(content):
            cls_name = m.group(1)
            # Capture inheritance from the declaration line
            decl_line = content[m.start() : m.start(2)]
            base_re = re.compile(r":\s*(?:public|protected|private)\s+(\w+)")
            base_m = base_re.search(decl_line)
            base_name = base_m.group(1) if base_m else ""
            brace_start = m.start(2)  # position of '{'
            # Walk forward to find matching '}' (tracking brace depth)
            depth = 0
            pos = brace_start
            while pos < len(content):
                ch = content[pos]
                if ch == "{":
                    depth += 1
                elif ch == "}":
                    depth -= 1
                    if depth == 0:
                        break
                pos += 1
            if depth != 0:
                # Unmatched braces — skip
                continue
            classes.append((cls_name, base_name, brace_start, pos + 1))

        for cls_name, base_name, body_start, body_end in classes:
            # Extract class body text
            body = content[body_start:body_end]

            is_set = is_settings_class(cls_name)
            has_save = bool(save_re.search(body))
            has_load = bool(load_re.search(body))
            has_op_eq = bool(operator_eq_re.search(body))
            has_eq_method = bool(equals_method_re.search(body))
            has_bin = bool(save_to_binary_re.search(body)) and bool(
                load_from_binary_re.search(body)
            )
            has_json = bool(save_to_json_re.search(body)) and bool(load_from_json_re.search(body))
            is_poly = bool(poly_load_re.search(body)) if (has_bin or has_json) else False

            if not (
                is_set or has_save or has_load or has_op_eq or has_eq_method or has_bin or has_json
            ):
                continue

            if cls_name not in results:
                results[cls_name] = ClassInfo(
                    name=cls_name,
                    is_settings=is_set,
                    header_file=rel,
                )
            info = results[cls_name]
            if is_set:
                info.is_settings = True
            if has_save and has_load:
                info.has_save_load = True
            if has_op_eq:
                info.has_operator_eq = True
            if has_eq_method:
                info.has_equals_method = True
            if has_bin:
                info.has_file_io_binary = True
            if has_json:
                info.has_file_io_json = True
            if is_poly:
                info.file_io_polymorphic = True
            if base_name:
                info.base_classes.append(base_name)

    # Propagate has_operator_eq down the inheritance chain.
    # operator== is inherited — if a base class has it, all derived classes do too.
    changed = True
    while changed:
        changed = False
        for cls_name, info in results.items():
            if info.has_operator_eq:
                continue
            for base in info.base_classes:
                if base in results and results[base].has_operator_eq:
                    info.has_operator_eq = True
                    changed = True
                    break

    # ── Phase 1b: Harvest derived classes not yet in results ────────────
    # Some derived types don't redeclare save/load/operator== themselves,
    # so the initial scan missed them.  Scan headers again, looking for
    # classes that inherit from any class we already track.
    class_start_re2 = re.compile(
        r"(?:^|\n)\s*(?:class|struct)\s+(\w+)\s*:\s*(?:public|protected|private)\s+(\w+)",
        re.MULTILINE,
    )
    for hdr in sorted(INCLUDE_DIR.rglob("*.h")):
        rel = str(hdr.relative_to(ROOT))
        try:
            content = hdr.read_text(errors="ignore")
        except Exception:
            continue
        for m in class_start_re2.finditer(content):
            derived, base = m.group(1), m.group(2)
            if base not in results:
                continue
            if derived in results:
                continue
            results[derived] = ClassInfo(
                name=derived,
                header_file=rel,
                base_classes=[base],
            )

    # Propagate has_operator_eq and file IO through inheritance.
    # operator==, saveTo*, loadFrom* are all inherited by derived classes.
    for attr in (
        "has_operator_eq",
        "has_file_io_binary",
        "has_file_io_json",
        "file_io_polymorphic",
    ):
        changed = True
        while changed:
            changed = False
            for cls_name, info in results.items():
                if getattr(info, attr):
                    continue
                for base in info.base_classes:
                    if base in results and getattr(results[base], attr):
                        setattr(info, attr, True)
                        changed = True
                        break

    return results


# ── Phase 2: C++ tests ────────────────────────────────────────────────────


def scan_cpp_tests(class_infos: dict[str, ClassInfo]) -> None:
    """Map serialization test files to the classes they test."""
    if not TESTS_CPP_DIR.exists():
        return

    for test_file in sorted(TESTS_CPP_DIR.glob("*.cpp")):
        if "serial" not in test_file.name.lower():
            continue
        try:
            content = test_file.read_text(errors="ignore")
        except Exception:
            continue
        rel = str(test_file.relative_to(ROOT))

        for cls_name, info in class_infos.items():
            if not info.has_save_load:
                continue
            base = strip_template(cls_name)
            if re.search(rf"\b{re.escape(base)}\b", content):
                info.cpp_test_files.append(rel)


# ── Phase 3: Python exposure ──────────────────────────────────────────────


def _find_py_bindings(content: str) -> list[tuple[str, int, int]]:
    """Return [(short_class_name, start, template_end), ...] for py::class_<...>."""
    results = []
    for m in re.finditer(r"py::class_\s*<", content):
        pos = m.end()
        depth = 0
        end = pos
        while end < len(content):
            c = content[end]
            if c == "<":
                depth += 1
            elif c == ">":
                if depth == 0:
                    break
                depth -= 1
            end += 1
        template = content[pos:end]

        # Get first argument (split on commas at depth 0)
        first_arg = ""
        d = 0
        for c in template:
            if c == "," and d == 0:
                break
            first_arg += c
            if c == "<":
                d += 1
            elif c == ">":
                d -= 1
        first_arg = first_arg.strip()

        # class name: last identifier before '<' or end
        lt = first_arg.find("<")
        name_part = first_arg[:lt] if lt >= 0 else first_arg
        nm = re.search(r"(\w+)\s*$", name_part)
        if nm:
            results.append((nm.group(1), m.start(), end))
    return results


def scan_python_exposure(class_infos: dict[str, ClassInfo]) -> None:
    """Detect __eq__, pickle, save_to in expose*.cpp files."""
    if not SRC_TUDATPY_DIR.exists():
        return

    # Build lookup: short_name → [cls_name, ...]
    name_map: dict[str, list[str]] = defaultdict(list)
    for cls_name in class_infos:
        name_map[strip_template(cls_name)].append(cls_name)

    for exp_file in sorted(SRC_TUDATPY_DIR.rglob("expose*.cpp")):
        try:
            content = exp_file.read_text(errors="ignore")
        except Exception:
            continue
        rel = str(exp_file.relative_to(ROOT))

        bindings = _find_py_bindings(content)

        for i, (short_name, _start, template_end) in enumerate(bindings):
            if short_name not in name_map:
                continue

            # Segment from template end to next binding start (or EOF)
            if i + 1 < len(bindings):
                seg_end = bindings[i + 1][1]
            else:
                seg_end = len(content)
            segment = content[template_end:seg_end]

            has_eq = bool(
                re.search(r'"__eq__"', segment)
                or re.search(r"py::self\s*==", segment)
                or re.search(r"TUDATPY_DEF_EQ_NE\s*\(", segment)
            )
            has_pickle = (
                "py::pickle" in segment
                or "make_pickle" in segment
                or bool(re.search(r"TUDATPY_DEF_PICKLE", segment))
            )
            has_save = bool(
                re.search(r'"save_(?:to_)?(?:json|binary|bin)"', segment)
                or re.search(r'"load_(?:from_)?(?:json|binary|bin)"', segment)
                or re.search(
                    r"TUDATPY_DEF_(?:FILE_IO(?:_POLYMORPHIC)?|BINARY_IO(?:_POLYMORPHIC)?|JSON_IO(?:_POLYMORPHIC)?)\s*\(",
                    segment,
                )
            )

            for cls_name in name_map[short_name]:
                info = class_infos[cls_name]
                info.py_expose_file = rel
                if has_eq:
                    info.py_has_equals = True
                if has_pickle:
                    info.py_has_pickle = True
                if has_save:
                    info.py_has_save_to = True


# ── Phase 4: Python tests ─────────────────────────────────────────────────


def scan_python_tests(class_infos: dict[str, ClassInfo]) -> None:
    """Detect Python roundtrip tests."""
    if not TESTS_PY_DIR.exists():
        return

    for py_test in sorted(TESTS_PY_DIR.rglob("*.py")):
        try:
            content = py_test.read_text(errors="ignore")
        except Exception:
            continue
        lower = content.lower()
        if "pickle" not in lower and "serial" not in lower:
            continue

        rel = str(py_test.relative_to(ROOT))

        for cls_name, info in class_infos.items():
            base = strip_template(cls_name)
            if re.search(rf"\b{re.escape(base)}\b", content):
                info.py_test_files.append(rel)
                continue
            # Also search for Python factory function names that create this class
            factories = CPP_TO_PYTHON_FACTORY.get(cls_name, [])
            for factory in factories:
                if re.search(rf"\b{re.escape(factory)}\b", content):
                    info.py_test_files.append(rel)
                    break


# ── Excel / CSV output ────────────────────────────────────────────────────


def _yesno(b: bool) -> str:
    return "✓" if b else "✗"


def write_excel(class_infos: dict[str, ClassInfo], output_path: Path) -> Path:
    try:
        import openpyxl
    except ImportError:
        print("openpyxl not installed. Install with: pip install openpyxl")
        return write_csv(class_infos, output_path.with_suffix(".csv"))

    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    hdr_font = Font(bold=True, size=11, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="4472C4")
    yes_fill = PatternFill("solid", fgColor="C6EFCE")
    no_fill = PatternFill("solid", fgColor="FFC7CE")
    settings_fill = PatternFill("solid", fgColor="FFEB9C")
    center = Alignment(horizontal="center", vertical="top")
    wrap = Alignment(wrap_text=True, vertical="top")

    # ── Sheet 1: All Classes ──
    ws = wb.active
    ws.title = "Serialization Status"

    headers = [
        "Class",
        "Is Settings",
        "Has Cereal\n(save/load)",
        "operator==",
        "equals() method",
        "File IO\nBinary",
        "File IO\nJSON",
        "File IO\nPolymorphic",
        "C++ Roundtrip\nTest",
        "Python Exposure\nFile",
        "Py: operator==",
        "Py: pickle",
        "Py: file I/O\n(save/load)",
        "Python\nRoundtrip Test",
        "Header File",
        "C++ Test Files",
        "Python Test Files",
    ]

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = wrap

    row = 2
    for cls_name in sorted(class_infos.keys()):
        info = class_infos[cls_name]
        vals = [
            cls_name,
            _yesno(info.is_settings),
            _yesno(info.has_save_load),
            _yesno(info.has_operator_eq),
            _yesno(info.has_equals_method),
            _yesno(info.has_file_io_binary),
            _yesno(info.has_file_io_json),
            _yesno(info.file_io_polymorphic),
            _yesno(bool(info.cpp_test_files)),
            info.py_expose_file or "",
            _yesno(info.py_has_equals),
            _yesno(info.py_has_pickle),
            _yesno(info.py_has_save_to),
            _yesno(bool(info.py_test_files)),
            info.header_file,
            "; ".join(info.cpp_test_files),
            "; ".join(info.py_test_files),
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.alignment = wrap
            center_cols = {2, 3, 4, 5, 6, 7, 8, 9, 11, 12, 13, 14}
            if c in center_cols:
                cell.alignment = center
                if v == "✓":
                    cell.fill = yes_fill
                elif v == "✗":
                    cell.fill = no_fill
        if info.is_settings:
            ws.cell(row=row, column=1).fill = settings_fill
        row += 1

    col_widths = [40, 12, 12, 12, 16, 14, 14, 16, 18, 50, 14, 14, 18, 18, 55, 55, 55]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}{row - 1}"

    # ── Sheet 2: Summary ──
    ws2 = wb.create_sheet("Summary")
    total = len(class_infos)
    settings = sum(1 for i in class_infos.values() if i.is_settings)

    def _c(attr, so=False):
        if so:
            return sum(1 for i in class_infos.values() if i.is_settings and getattr(i, attr))
        return sum(1 for i in class_infos.values() if getattr(i, attr))

    def _cb(attr, so=False):
        if so:
            return sum(1 for i in class_infos.values() if i.is_settings and bool(getattr(i, attr)))
        return sum(1 for i in class_infos.values() if bool(getattr(i, attr)))

    for r, row_data in enumerate(
        [
            ("Metric", "All Classes", "Settings Only"),
            ("Total", total, settings),
            ("Has cereal save/load", _c("has_save_load"), _c("has_save_load", True)),
            ("Has operator==", _c("has_operator_eq"), _c("has_operator_eq", True)),
            ("Has equals() method", _c("has_equals_method"), _c("has_equals_method", True)),
            ("Has File IO Binary", _c("has_file_io_binary"), _c("has_file_io_binary", True)),
            ("Has File IO JSON", _c("has_file_io_json"), _c("has_file_io_json", True)),
            ("File IO Polymorphic", _c("file_io_polymorphic"), _c("file_io_polymorphic", True)),
            ("C++ roundtrip test", _cb("cpp_test_files"), _cb("cpp_test_files", True)),
            ("Py Binding: operator==", _c("py_has_equals"), _c("py_has_equals", True)),
            ("Py Binding: pickle", _c("py_has_pickle"), _c("py_has_pickle", True)),
            ("Py Binding: file I/O", _c("py_has_save_to"), _c("py_has_save_to", True)),
            ("Python roundtrip test", _cb("py_test_files"), _cb("py_test_files", True)),
        ],
        1,
    ):
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            if r == 1:
                cell.font = hdr_font
                cell.fill = hdr_fill
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 16
    ws2.column_dimensions["C"].width = 16

    # ── Sheet 3: Settings classes ──
    ws3 = wb.create_sheet("Settings Classes")
    s_headers = [
        "Class",
        "Save/Load",
        "operator==",
        "equals() method",
        "File IO\nBinary",
        "File IO\nJSON",
        "File IO\nPolymorphic",
        "C++ Test",
        "Py: operator==",
        "Py: pickle",
        "Py: file I/O",
        "Py Test",
        "Header File",
    ]
    for c, h in enumerate(s_headers, 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill

    row = 2
    for cls_name in sorted(class_infos.keys()):
        info = class_infos[cls_name]
        if not info.is_settings:
            continue
        vals = [
            cls_name,
            _yesno(info.has_save_load),
            _yesno(info.has_operator_eq),
            _yesno(info.has_equals_method),
            _yesno(info.has_file_io_binary),
            _yesno(info.has_file_io_json),
            _yesno(info.file_io_polymorphic),
            _yesno(bool(info.cpp_test_files)),
            _yesno(info.py_has_equals),
            _yesno(info.py_has_pickle),
            _yesno(info.py_has_save_to),
            _yesno(bool(info.py_test_files)),
            info.header_file,
        ]
        for c, v in enumerate(vals, 1):
            cell = ws3.cell(row=row, column=c, value=v)
            center_cols = {2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12}
            if c in center_cols:
                cell.alignment = center
                if v == "✓":
                    cell.fill = yes_fill
                elif v == "✗":
                    cell.fill = no_fill
        row += 1

    for i, w in enumerate([45, 12, 14, 16, 14, 14, 16, 12, 14, 14, 18, 12, 60], 1):
        ws3.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(s_headers))}{row - 1}"

    wb.save(output_path)
    return output_path


def write_csv(class_infos: dict[str, ClassInfo], output_path: Path) -> Path:
    import csv

    rows = []
    for cls_name in sorted(class_infos.keys()):
        info = class_infos[cls_name]
        rows.append(
            {
                "Class": cls_name,
                "Is Settings": "yes" if info.is_settings else "no",
                "Save/Load": "yes" if info.has_save_load else "no",
                "operator==": "yes" if info.has_operator_eq else "no",
                "equals() method": "yes" if info.has_equals_method else "no",
                "File IO Binary": "yes" if info.has_file_io_binary else "no",
                "File IO JSON": "yes" if info.has_file_io_json else "no",
                "File IO Polymorphic": "yes" if info.file_io_polymorphic else "no",
                "C++ Roundtrip Test": "yes" if info.cpp_test_files else "no",
                "Python Exposure File": info.py_expose_file,
                "Py Binding: operator==": "yes" if info.py_has_equals else "no",
                "Py Binding: pickle": "yes" if info.py_has_pickle else "no",
                "Py Binding: file I/O": "yes" if info.py_has_save_to else "no",
                "Python Roundtrip Test": "yes" if info.py_test_files else "no",
                "Header File": info.header_file,
                "C++ Test Files": "; ".join(info.cpp_test_files),
                "Python Test Files": "; ".join(info.py_test_files),
            }
        )

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        if not rows:
            f.write("No classes found\n")
            return output_path
        w = csv.DictWriter(f, fieldnames=rows[0].keys())
        w.writeheader()
        w.writerows(rows)
    return output_path


# ── main ────────────────────────────────────────────────────────────────────


def main():
    parser = argparse.ArgumentParser(
        description="Track serialization compliance across the Tudat codebase."
    )
    parser.add_argument("--csv", action="store_true", help="Output CSV instead of Excel")
    parser.add_argument(
        "--output", "-o", type=Path, default=None, help=f"Output path (default: {DEFAULT_EXCEL})"
    )
    args = parser.parse_args()

    print("Scanning headers for serializable classes...")
    class_infos = scan_headers()
    print(f"  Found {len(class_infos)} classes")

    print("Scanning C++ test files...")
    scan_cpp_tests(class_infos)

    print("Scanning Python exposure files...")
    scan_python_exposure(class_infos)

    # Populate Python factory names from the static mapping table.
    # This allows the tracker to detect Python test coverage even when the
    # test code uses factory function names that differ from C++ class names.
    print("Populating Python factory name mappings...")
    for cls_name, info in class_infos.items():
        info.py_factory_names = CPP_TO_PYTHON_FACTORY.get(cls_name, [])

    print("Scanning Python test files...")
    scan_python_tests(class_infos)

    # Propagate Python exposure status down the inheritance chain.
    # A bound base with __eq__/pickle/save_to exposes those features for all
    # derived types through polymorphic serialization — even if the derived
    # type has no dedicated py::class_ of its own.
    print("Propagating Python exposure through inheritance...")
    for attr in ("py_has_equals", "py_has_pickle", "py_has_save_to", "py_expose_file"):
        changed = True
        while changed:
            changed = False
            for cls_name, info in class_infos.items():
                if getattr(info, attr):
                    continue
                for base in info.base_classes:
                    if base in class_infos and getattr(class_infos[base], attr):
                        setattr(info, attr, getattr(class_infos[base], attr))
                        changed = True
                        break

    # Output
    if args.csv:
        path = write_csv(class_infos, args.output or DEFAULT_EXCEL.with_suffix(".csv"))
    else:
        path = write_excel(class_infos, args.output or DEFAULT_EXCEL)

    print(f"\nDone. Report written to: {path}")

    # Quick summary
    total = len(class_infos)
    sl = sum(1 for i in class_infos.values() if i.has_save_load)
    op_eq = sum(1 for i in class_infos.values() if i.has_operator_eq)
    eq_method = sum(1 for i in class_infos.values() if i.has_equals_method)
    cpp = sum(1 for i in class_infos.values() if i.cpp_test_files)
    py_eq = sum(1 for i in class_infos.values() if i.py_has_equals)
    py_pk = sum(1 for i in class_infos.values() if i.py_has_pickle)
    py_st = sum(1 for i in class_infos.values() if i.py_has_save_to)
    py_t = sum(1 for i in class_infos.values() if i.py_test_files)
    settings = sum(1 for i in class_infos.values() if i.is_settings)
    file_io_bin = sum(1 for i in class_infos.values() if i.has_file_io_binary)
    file_io_json = sum(1 for i in class_infos.values() if i.has_file_io_json)
    file_io_poly = sum(1 for i in class_infos.values() if i.file_io_polymorphic)
    print(f"\n{'='*68}")
    print(f"  Total: {total}  |  Settings: {settings}")
    print(
        f"  save/load: {sl}  |  operator==: {op_eq}  |  equals(): {eq_method}  |  C++ tests: {cpp}"
    )
    print(
        f"  File IO binary: {file_io_bin}  |  File IO JSON: {file_io_json}  |  Polymorphic: {file_io_poly}"
    )
    print(
        f"  Py Binding: operator==: {py_eq}  |  Py Binding: pickle: {py_pk}  |  Py Binding: file I/O: {py_st}"
    )
    print(f"  Py tests:  {py_t}")
    print(f"{'='*68}")


if __name__ == "__main__":
    main()
